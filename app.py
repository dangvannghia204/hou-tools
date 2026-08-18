import os
import tempfile
import shutil
import re
import math
import zipfile
from datetime import datetime
from copy import copy
import pandas as pd
import numpy as np
import polars as pl
from python_calamine import CalamineWorkbook
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import streamlit as st

try:
    from pyxlsb import open_workbook as open_xlsb
except ImportError:
    pass

# ==========================================
# CẤU HÌNH TRANG & CSS COMPACT (SINGLE-SCREEN)
# ==========================================
st.set_page_config(page_title="Excel Data Workspace", page_icon="📊", layout="wide")

st.markdown("""
    <style>
        .block-container { 
            padding-top: 1rem !important; 
            padding-bottom: 0rem !important; 
            max-width: 1250px;
        }
        .app-header {
            background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
            padding: 1rem 1.5rem;
            border-radius: 8px;
            color: white;
            margin-bottom: 1rem;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
        }
        .app-header h2 { color: white; margin: 0; font-size: 1.6rem; font-weight: 700; padding-bottom: 0.2rem;}
        .app-header p { margin: 0; font-size: 0.95rem; opacity: 0.9; }

        .instruction-card {
            background-color: #F8FAFC;
            border-left: 4px solid #3B82F6;
            padding: 1rem;
            border-radius: 6px;
            color: #334155;
            font-size: 0.95rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            margin-bottom: 1rem;
            min-height: 160px;
        }
        
        [data-testid="stFileUploadDropzone"] {
            min-height: 100px !important;
            padding: 1rem !important;
        }
        [data-testid="stFileUploadDropzone"] div { gap: 0.2rem; }
        
        .stButton>button {
            width: 100%;
            border-radius: 6px;
            height: 45px;
            font-size: 15px;
            font-weight: 600;
            background-color: #2563EB;
            color: white;
            border: none;
            transition: all 0.2s ease-in-out;
        }
        .stButton>button:hover { background-color: #1D4ED8; transform: translateY(-1px); }
        
        .stDownloadButton>button {
            width: 100%;
            border-radius: 6px;
            height: 45px;
            font-size: 15px;
            font-weight: 600;
            background-color: #10B981; 
            color: white;
            border: none;
            transition: all 0.2s ease-in-out;
        }
        .stDownloadButton>button:hover { background-color: #059669; transform: translateY(-1px); }
        
        css-1v0mbdj { margin-top: 0rem; }
        .st-emotion-cache-1kyxreq { gap: 0.5rem; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# CÁC CLASS & HÀM XỬ LÝ LÕI
# ==========================================
class ExcelDataProcessor:
    def __init__(self, input_file="Data_fill.xlsx", keywords_str="DNP, HCM", log_callback=print):
        self.input_file = input_file
        self.keywords_str = keywords_str
        self.log = log_callback
        self.base_dir = os.path.dirname(os.path.abspath(input_file)) if os.path.dirname(input_file) else os.getcwd()

    def run_all_steps(self):
        if not os.path.exists(self.input_file): return False
        self._preprocess_khlm()
        out_khl = os.path.join(self.base_dir, "Ketqua_KHLM.xlsx")
        self._run_source_1(is_hoc_lai=False, output_path=out_khl)
        out_hl = os.path.join(self.base_dir, "Ketquahoclai_KHLM.xlsx")
        self._run_source_1(is_hoc_lai=True, output_path=out_hl)
        out_merge = os.path.join(self.base_dir, "Data_Merge.xlsx")
        self._merge_results(out_khl, out_hl, out_merge)
        out_lopmon = os.path.join(self.base_dir, "LopMon_Ketqua.xlsx")
        self._run_source_2_and_map(out_merge, out_lopmon)
        return True

    def _preprocess_khlm(self):
        wb = load_workbook(self.input_file)
        actual_khlm = next((s for s in wb.sheetnames if s.lower() == 'khlm'), None)
        if not actual_khlm: return

        df_khlm = pd.read_excel(self.input_file, sheet_name=actual_khlm).astype(object)
        cols = df_khlm.columns.tolist()
        col_map = {str(c).strip().upper(): c for c in cols}
        keywords = [k.strip().upper() for k in self.keywords_str.split(',') if k.strip()]

        for index, row in df_khlm.iterrows():
            ten_lop = str(row.get('TenLop', '')).upper()
            dia_phuong = str(row.get('DiaPhuong', ''))
            dia_phuong_upper = dia_phuong.upper()

            if 'HL' in ten_lop or 'HL' in dia_phuong_upper or 'HỌC LẠI' in dia_phuong_upper:
                df_khlm.at[index, 'DiaPhuongHL'] = 'HL'

            current_dp_khl = str(row.get('DiaPhuongKHL', ''))
            if current_dp_khl.lower() == 'nan': current_dp_khl = ''
            
            added_codes = []
            for kw in keywords:
                if kw in dia_phuong_upper:
                    if kw in col_map: 
                        original_col_name = col_map[kw]
                        val = row[original_col_name]
                        if pd.notna(val) and isinstance(val, (int, float)) and val > 0:
                            added_codes.append(kw)
            
            if added_codes:
                str_to_add = ",".join(added_codes)
                if current_dp_khl.strip(): df_khlm.at[index, 'DiaPhuongKHL'] = f"{current_dp_khl},{str_to_add}"
                else: df_khlm.at[index, 'DiaPhuongKHL'] = str_to_add

        with pd.ExcelWriter(self.input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
             df_khlm.to_excel(writer, sheet_name=actual_khlm, index=False)

    def _get_col_idx(self, df, col_name, default_name):
        if col_name in df.columns: return df.columns.get_loc(col_name)
        else: return -1

    def _run_source_1(self, is_hoc_lai, output_path):
        xls = pd.ExcelFile(self.input_file)
        actual_khlm = next((s for s in xls.sheet_names if s.lower() == 'khlm'), None)
        actual_data = next((s for s in xls.sheet_names if s.lower() == 'data'), None)
        if not actual_khlm or not actual_data: return

        khlm = pd.read_excel(self.input_file, sheet_name=actual_khlm, header=0).astype(object)
        data = pd.read_excel(self.input_file, sheet_name=actual_data)

        idx_tenlop = self._get_col_idx(khlm, 'TenLop', 'TenLop')
        idx_mamon_khlm = self._get_col_idx(khlm, 'MaMon', 'MaMon')
        idx_dp_khl = self._get_col_idx(khlm, 'DiaPhuongKHL', 'DiaPhuongKHL')
        idx_dp_hl = self._get_col_idx(khlm, 'DiaPhuongHL', 'DiaPhuongHL')

        idx_loplt_data = self._get_col_idx(data, 'LopLT', 'LopLT')
        idx_mamon_data = self._get_col_idx(data, 'MaMon', 'MaMon')
        idx_matram_data = self._get_col_idx(data, 'MaTram', 'MaTram')
        idx_msv_data = self._get_col_idx(data, 'MSV', 'MSV')

        assigned = {}
        total_rows = len(khlm)
        filter_keywords = [k.strip().upper() for k in self.keywords_str.split(',') if k.strip()]
        n = idx_dp_hl if is_hoc_lai else idx_dp_khl

        for i in range(total_rows):
            row = khlm.iloc[i]
            kc = str(row.iloc[idx_mamon_khlm]).strip().upper()
            kh = str(row.iloc[idx_tenlop]).strip().upper()
            kj = str(row.iloc[n]).strip().upper()
            
            if kc == 'NAN' or not kc or kc == "": continue

            kc_parts = [p.strip() for p in kc.split('/') if p.strip()]
            mask_mamon = data.iloc[:, idx_mamon_data].astype(str).str.upper().apply(
                lambda x: any((x.strip() in p) or (p in x.strip()) for p in kc_parts)
            )
            mask_matram = data.iloc[:, idx_matram_data].astype(str).str.upper().apply(
                lambda x: (x.strip() in kj) or (kj in x.strip())
            )

            if not is_hoc_lai:
                mask_loplt = data.iloc[:, idx_loplt_data].astype(str).str.upper().apply(
                    lambda x: (x.strip() in kh) or (kh in x.strip())
                )
                mask = mask_mamon & mask_loplt & mask_matram
            else: mask = mask_mamon & mask_matram

            msv_col_name = data.columns[idx_msv_data]
            matched_msvs = data.loc[mask, msv_col_name].unique().tolist()

            if any(kw in kj for kw in filter_keywords):
                if kc not in assigned: assigned[kc] = set()
                final_msvs = [m for m in matched_msvs if m not in assigned[kc]]
                assigned[kc].update(final_msvs)
                matched_msvs = final_msvs

            khlm.iloc[i, 4] = ", ".join(map(str, matched_msvs))
            khlm.iloc[i, 5] = len(matched_msvs)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            khlm.to_excel(writer, sheet_name='Result', index=False)

    def _merge_results(self, file_khl, file_hl, file_merge_out):
        xls1 = pd.ExcelFile(file_khl)
        actual_res1 = next((s for s in xls1.sheet_names if s.lower() == 'result'), None)
        xls2 = pd.ExcelFile(file_hl)
        actual_res2 = next((s for s in xls2.sheet_names if s.lower() == 'result'), None)
        
        if not actual_res1 or not actual_res2: return

        df1 = pd.read_excel(file_khl, sheet_name=actual_res1, usecols=[0, 1, 2, 3, 4])
        df2 = pd.read_excel(file_hl, sheet_name=actual_res2, usecols=[0, 1, 2, 3, 4])

        df1_valid = df1[df1[df1.columns[4]].astype(str).str.strip().replace('nan', '') != '']
        df2_valid = df2[df2[df2.columns[4]].astype(str).str.strip().replace('nan', '') != '']

        df_merged = pd.concat([df1_valid, df2_valid], ignore_index=True)
        df_merged.to_excel(file_merge_out, sheet_name='data', index=False)

    def _run_source_2_and_map(self, file_merge_in, file_lopmon_out):
        wb = load_workbook(file_merge_in)
        data_sheet_name = next((s for s in wb.sheetnames if s.lower() == 'data'), None)
        if not data_sheet_name: return
        ws_data = wb[data_sheet_name]
        
        actual_res = next((s for s in wb.sheetnames if s.lower() == 'result'), None)
        if actual_res: del wb[actual_res]
        ws_result = wb.create_sheet('Result')
        ws_result.append(["Cột A (M Mã tách)", "Cột B (Từ A)", "Cột C (Từ B)", "Cột D (Từ C)", "Cột E (Từ D)", "Cột F (Trống)", "Cột G (A+D)", "Cột H (Copy B)"])
        result_data_for_mapping = {}

        for row in range(2, ws_data.max_row + 1):
            val_a = ws_data.cell(row=row, column=1).value
            val_b = ws_data.cell(row=row, column=2).value
            val_c = ws_data.cell(row=row, column=3).value
            val_d = ws_data.cell(row=row, column=4).value
            val_e = ws_data.cell(row=row, column=5).value

            if val_a is None and val_b is None and val_e is None: continue

            if val_e is not None:
                for item in str(val_e).split(','):
                    clean_item = item.strip()
                    if clean_item:
                        col_g = str(clean_item) + str(val_c if val_c is not None else "")
                        col_h = val_a
                        ws_result.append([clean_item, val_a, val_b, val_c, val_d, "", col_g, col_h])
                        result_data_for_mapping[col_g] = col_h
                        if val_c and '/' in str(val_c):
                            for part in str(val_c).split('/'):
                                part_clean = part.strip()
                                if part_clean: result_data_for_mapping[str(clean_item) + part_clean] = col_h
            else:
                 col_g = "" + str(val_c if val_c is not None else "")
                 col_h = val_a
                 ws_result.append(["", val_a, val_b, val_c, val_d, "", col_g, col_h])
                 result_data_for_mapping[col_g] = col_h
                 if val_c and '/' in str(val_c):
                     for part in str(val_c).split('/'):
                         part_clean = part.strip()
                         if part_clean: result_data_for_mapping["" + part_clean] = col_h
        wb.save(file_lopmon_out)

        wb_fill = load_workbook(self.input_file)
        actual_fill_data = next((s for s in wb_fill.sheetnames if s.lower() == 'data'), None)
        if not actual_fill_data: return
        ws_fill_data = wb_fill[actual_fill_data]
        
        header_row = 1
        col_idx_msv, col_idx_mamon, col_idx_macoursett = -1, -1, -1
        for cell in ws_fill_data[header_row]:
            val = str(cell.value).strip().upper() if cell.value else ""
            if val == 'MSV': col_idx_msv = cell.column
            elif val == 'MAMON': col_idx_mamon = cell.column
            elif val == 'MACOURSETT': col_idx_macoursett = cell.column

        if col_idx_msv == -1 or col_idx_mamon == -1: return
        if col_idx_macoursett == -1:
            col_idx_macoursett = ws_fill_data.max_column + 1
            ws_fill_data.cell(row=header_row, column=col_idx_macoursett, value="MaCourseTT")

        for row in range(2, ws_fill_data.max_row + 1):
            msv = str(ws_fill_data.cell(row=row, column=col_idx_msv).value or "").strip()
            mamon = str(ws_fill_data.cell(row=row, column=col_idx_mamon).value or "").strip()
            lookup_key = msv + mamon 
            if lookup_key in result_data_for_mapping:
                ws_fill_data.cell(row=row, column=col_idx_macoursett, value=result_data_for_mapping[lookup_key])
            elif '/' in mamon:
                for part in mamon.split('/'):
                    part_clean = part.strip()
                    alt_key = msv + part_clean
                    if alt_key in result_data_for_mapping:
                        ws_fill_data.cell(row=row, column=col_idx_macoursett, value=result_data_for_mapping[alt_key])
                        break 
        wb_fill.save(os.path.join(self.base_dir, "Data_fill_Finish.xlsx"))

# ==========================================
# CÁC HÀM TIỆN ÍCH CHUNG
# ==========================================
def apply_full_border(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row: cell.border = border

def auto_fit_columns(ws, padding=2):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length: max_length = length
            except: pass
        ws.column_dimensions[column].width = max(10, min(max_length + padding, 50))

def roman_to_int(r):
    m = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10}
    return m.get(str(r).upper(), r)

def format_hk(a6):
    if not a6: return ""
    try:
        rp = re.search(r'Học kỳ\s+([I|V|X]+)', str(a6), re.IGNORECASE)
        hn = roman_to_int(rp.group(1)) if rp else ""
        yp = re.search(r'(\d{4})-(\d{4})', str(a6))
        if yp: return f"HK{hn}_{yp.group(1)[2:]}.{yp.group(2)[2:]}"
        return f"HK{hn}"
    except: return str(a6)

def format_excel_date(cell):
    v = cell.value
    return v.strftime("%d/%m/%Y") if isinstance(v, datetime) else (str(v) if v else "")

# ==========================================
# CÁC HÀM XỬ LÝ (WRAPPERS)
# ==========================================
def check_tien_quyet_logic(folder_path):
    ketqua_file = None
    dktq_file = None
    for f in os.listdir(folder_path):
        if f.lower() == 'ketqua.xlsx':
            ketqua_file = os.path.join(folder_path, f)
        elif f.lower() == 'dktq.xlsx':
            dktq_file = os.path.join(folder_path, f)

    if not ketqua_file or not dktq_file:
        raise ValueError("Không tìm thấy đủ 2 file 'Ketqua.xlsx' và 'dktq.xlsx' trong thư mục đã tải lên!")

    wb_kq = load_workbook(ketqua_file)
    ws_kq = wb_kq.active
    b_groups = {}
    for r in range(2, ws_kq.max_row + 1):
        cell_d = ws_kq.cell(row=r, column=4).value
        cell_b = ws_kq.cell(row=r, column=2).value
        m_val = ""
        if cell_d is not None:
            d_str = str(cell_d)
            if '.' in d_str: m_val = d_str.rsplit('.', 1)[0]
            else: m_val = d_str
            ws_kq.cell(row=r, column=13).value = m_val

        if cell_b is not None:
            b_str = str(cell_b).strip()
            if b_str not in b_groups:
                b_groups[b_str] = {'first_row': r, 'm_vals': []}
            if m_val:
                b_groups[b_str]['m_vals'].append(m_val)

    n_dict = {}
    for b_str, data in b_groups.items():
        m_list = data['m_vals']
        if m_list: n_val_str = f"{b_str}," + ",".join(m_list)
        else: n_val_str = b_str
        ws_kq.cell(row=data['first_row'], column=14).value = n_val_str
        n_dict[b_str] = {x.strip() for x in n_val_str.split(',') if x.strip()}

    out_kq = os.path.join(folder_path, "Ketqua_Finish.xlsx")
    wb_kq.save(out_kq)

    wb_dktq = load_workbook(dktq_file)
    ws_dktq = wb_dktq.active
    keyword = "Môn tiên quyết chưa hoàn thành "

    for r in range(2, ws_dktq.max_row + 1):
        cell_m = ws_dktq.cell(row=r, column=13).value
        cell_d = ws_dktq.cell(row=r, column=4).value
        d_str = str(cell_d).strip() if cell_d is not None else ""
        extracted_m = ""

        if cell_m is not None:
            m_str = str(cell_m)
            if keyword in m_str:
                extracted_m = m_str.split(keyword, 1)[1]

        if extracted_m:
            t_val_str = f"{d_str},{extracted_m}" if d_str else extracted_m
        else:
            t_val_str = d_str

        ws_dktq.cell(row=r, column=20).value = t_val_str
        t_set = {x.strip() for x in t_val_str.split(',') if x.strip()}
        target_n_set = n_dict.get(d_str, set())

        if t_set and t_set.issubset(target_n_set):
            ws_dktq.cell(row=r, column=21).value = "CPĐK"
        else:
            missing_elements = t_set - target_n_set
            if missing_elements:
                ws_dktq.cell(row=r, column=21).value = ", ".join(sorted(list(missing_elements)))
            else:
                ws_dktq.cell(row=r, column=21).value = ""

    out_dktq = os.path.join(folder_path, "dktq_Finish.xlsx")
    wb_dktq.save(out_dktq)

    zip_path = os.path.join(folder_path, "KetQua_KiemTra_TienQuyet.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        zipf.write(out_kq, "Ketqua_Finish.xlsx")
        zipf.write(out_dktq, "dktq_Finish.xlsx")

    return zip_path

def clean_value_gom_diem(val):
    if val is None: return ""
    if isinstance(val, float):
         if math.isnan(val): return ""
         if val.is_integer(): return str(int(val))
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): return s[:-2]
    return s

def extract_meta_gom_diem(df_str, row_idx, col_idx):
    try:
        if row_idx < df_str.height and col_idx < df_str.width:
            val = str(df_str.item(row_idx, col_idx))
            if ":" in val: return val.split(":", 1)[1].strip()
    except Exception: pass
    return ""

def extract_meta_multi_gom_diem(df_str, row_idx, col_indices):
    for col_idx in col_indices:
        try:
            if row_idx < df_str.height and col_idx < df_str.width:
                val = str(df_str.item(row_idx, col_idx))
                if ":" in val: return val.split(":", 1)[1].strip()
        except Exception: pass
    return ""

def gom_diem_logic(msv_path, data_dir):
    wb_msv = CalamineWorkbook.from_path(msv_path)
    sheet_msv = wb_msv.get_sheet_by_index(0)
    try: raw_msv = sheet_msv.to_python(skip_empty_area=False)
    except TypeError: raw_msv = sheet_msv.to_python()
    
    if not raw_msv or len(raw_msv) < 2:
        raise ValueError("File MSV rỗng.")
        
    headers = [str(x).strip().upper() if x is not None else "" for x in raw_msv[0]]
    try:
        tksv_idx = headers.index('TKSV')
    except ValueError:
        raise ValueError("Không tìm thấy cột 'TKSV' trong file đầu vào.")
        
    raw_ids = [clean_value_gom_diem(row[tksv_idx]) for row in raw_msv[1:] if len(row) > tksv_idx]
    student_ids_to_find = {x for x in raw_ids if x != ""}
    if not student_ids_to_find:
        raise ValueError("Cột 'TKSV' không có dữ liệu hợp lệ.")
    
    results_list = []
    search_list = list(student_ids_to_find)
    final_order = [
        'MSV Tìm thấy', 'Nguồn (File)', 'Sheet', 'TT', 'Họ và tên', 'Ngày sinh', 'Lớp',
        'Mã học phần', 'Tên học phần', 'Ngày thi', 'Ca thi', 'Đề số/ Mã đề',
        'Quá trình', 'Kiểm tra', 'Đề án', 'Thi', 'BTL', 'BCHC', 'Vấn đáp', 'TB', 'Tổng kết', 'Ghi chú'
    ]
    
    for dirpath, _, filenames in os.walk(data_dir):
        for filename in filenames:
            if filename.lower().endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')) and filename.lower() != 'result.xlsx' and not filename.startswith('~$'):
                filepath = os.path.join(dirpath, filename)
                try:
                    wb = CalamineWorkbook.from_path(filepath)
                    for sheet_name in wb.sheet_names:
                        sheet = wb.get_sheet_by_name(sheet_name)
                        try: raw_data = sheet.to_python(skip_empty_area=False)
                        except TypeError: raw_data = sheet.to_python()
                        
                        if not raw_data: continue
                        max_cols = max(len(row) for row in raw_data)
                        if max_cols == 0: continue
                            
                        str_data = []
                        for row in raw_data:
                            str_row = [clean_value_gom_diem(val) for val in row]
                            if len(str_row) < max_cols:
                                str_row.extend([""] * (max_cols - len(str_row)))
                            str_data.append(str_row)
                        
                        df_str = pl.DataFrame(str_data, orient="row")
                        a1_val = df_str.item(0, 0) if df_str.height > 0 and df_str.width > 0 else ""
                        a1_val_clean = ' '.join(str(a1_val).lower().split())
                        is_bgd = (a1_val_clean == "bộ giáo dục và đào tạo")
                        
                        ten_hp, ma_hp, ngay_thi, ca_thi = "", "", "", ""
                        max_row_idx = df_str.height - 1
                        
                        if is_bgd:
                            start_row_idx = 13
                            ten_hp = extract_meta_gom_diem(df_str, 5, 0)
                            ma_hp = extract_meta_gom_diem(df_str, 6, 0)
                            ngay_thi = extract_meta_multi_gom_diem(df_str, 6, [7, 8, 9])
                            ca_thi = extract_meta_multi_gom_diem(df_str, 7, [7, 8, 9])
                        else:
                            start_row_idx = 1

                        if df_str.height > start_row_idx:
                            df_search_area = df_str.slice(start_row_idx)
                            mask = pl.any_horizontal([pl.col(c).is_in(search_list) for c in df_search_area.columns])
                            matched_str_data = df_search_area.filter(mask)
                        else:
                            matched_str_data = pl.DataFrame()
                        
                        if not matched_str_data.is_empty():
                            header_map = {}
                            for col_idx_num, _ in enumerate(df_str.columns):
                                header_vals = []
                                if is_bgd:
                                    for r_idx in [10, 11, 12]:
                                        if r_idx <= max_row_idx:
                                            val = df_str.item(r_idx, col_idx_num)
                                            if val and str(val).lower() != 'nan' and str(val).strip():
                                                header_vals.append(' '.join(str(val).lower().split()))
                                else:
                                    if 0 <= max_row_idx:
                                        val = df_str.item(0, col_idx_num)
                                        if val and str(val).lower() != 'nan' and str(val).strip():
                                            header_vals.append(' '.join(str(val).lower().split()))
                                header_map[col_idx_num] = header_vals

                            for row_tuple in matched_str_data.iter_rows():
                                matched_id = "Không xác định"
                                for val in row_tuple:
                                    if val in student_ids_to_find:
                                        matched_id = val
                                        break

                                result_row = {col: "" for col in final_order}
                                result_row['MSV Tìm thấy'] = matched_id
                                result_row['Nguồn (File)'] = filename
                                result_row['Sheet'] = sheet_name
                                
                                if is_bgd:
                                    result_row['Mã học phần'] = ma_hp
                                    result_row['Tên học phần'] = ten_hp                                            
                                    result_row['Ngày thi'] = ngay_thi
                                    result_row['Ca thi'] = ca_thi

                                skip_next_col_for_name = False
                                for i in range(len(row_tuple)):
                                    if skip_next_col_for_name:
                                        skip_next_col_for_name = False
                                        continue

                                    val_to_add = row_tuple[i]
                                    raw_headers = header_map.get(i, [])
                                    if not raw_headers: continue 

                                    NAME_HEADERS = {'họ và tên', 'họ tên', 'họ và', 'họ', 'họ lót', 'họ đệm', 'tên'}
                                    matched_name_header = ""
                                    for rh in raw_headers:
                                        if rh in NAME_HEADERS:
                                            matched_name_header = rh
                                            break

                                    if matched_name_header:
                                        if val_to_add:
                                            val_to_add = str(val_to_add).strip()
                                            if matched_name_header != 'tên':
                                                result_row['Họ và tên'] = (result_row['Họ và tên'] + " " + val_to_add).strip()
                                            else:
                                                if val_to_add not in result_row['Họ và tên']:
                                                    result_row['Họ và tên'] = (result_row['Họ và tên'] + " " + val_to_add).strip()

                                        if i + 1 < len(row_tuple) and matched_name_header != 'tên':
                                            next_headers = header_map.get(i + 1, [])
                                            is_next_target = False
                                            if not next_headers: is_next_target = True
                                            else:
                                                for nh in next_headers:
                                                    if nh in {'tên'}:
                                                        is_next_target = True
                                                        break
                                            if is_next_target:
                                                next_val = row_tuple[i + 1]
                                                if next_val: result_row['Họ và tên'] = (result_row['Họ và tên'] + " " + str(next_val).strip()).strip()
                                                skip_next_col_for_name = True
                                        continue

                                    mapping_rules = {
                                        'TT': ['tt', 'stt'],
                                        'Ngày sinh': ['ngày sinh'],
                                        'Lớp': ['lớp'],
                                        'Mã học phần': ['mã học phần', 'mã môn', 'mã học phần (module)'],
                                        'Tên học phần': ['tên học phần', 'tên môn', 'tên học phần (module)'],                                                
                                        'Ngày thi': ['ngày thi'],
                                        'Ca thi': ['ca thi'],
                                        'Đề số/ Mã đề': ['đề số/ mã đề', 'đề số/mã đề', 'đề số', 'mã đề', 'đề thi'],
                                        'Quá trình': ['quá trình', 'đqt', 'điểm quá trình', 'điểm qt', 'qt'],
                                        'Kiểm tra': ['kiểm tra', 'đkt', 'kiểm tra(10%)', 'kiểm tra (10%)', 'điểm kiểm tra', 'điểm kt', 'kt'],
                                        'Đề án': ['đề án', 'đề án(70%)', 'đề án (70%)'],
                                        'Thi': ['thi', 'điểm thi'],
                                        'BTL': ['btl', 'bài tập lớn', 'điểm btl'],
                                        'BCHC': ['bchc', 'điểm bchc', 'điểmbchc', 'điểm bchc (đvới môn có điểm bchc trên ht)'],
                                        'Vấn đáp': ['vấn đáp', 'vđ', 'vđ(30%)', 'vđ (30%)', 'điểm vđ', 'điểm vấn đáp'],
                                        'TB': ['tb', 'tb (btl+vđ)/2', 'thi (đề án (70%)+vđ (30%)', 'thi(đềán(70%)+vđ(30%)', 'đề án+vđ', 'tb(btl+vđ)/2', 'tb(bchc+vđ)/2', 'tb(đề án+vđ)/2', 'trung bình', 'điểm tb'],
                                        'Tổng kết': ['điểm học phần', 'tổng kết', 'điểm tổng kết', 'điểm tk', 'điểm học phần (100%)'],
                                        'Ghi chú': ['ghi chú']
                                    }

                                    for std_col, keywords in mapping_rules.items():
                                        matched = False
                                        for rh in raw_headers:
                                            rh_no_space = rh.replace(" ", "")
                                            for kw in keywords:
                                                kw_no_space = kw.replace(" ", "")
                                                if rh == kw or rh_no_space == kw_no_space:
                                                    matched = True
                                                    break 
                                            if matched: break
                                        if matched:
                                            if result_row[std_col] == "": result_row[std_col] = val_to_add
                                            break
                                results_list.append(result_row)
                except Exception:
                    pass

    if results_list:
        merged_dict = {}
        for row in results_list:
            msv = row.get('MSV Tìm thấy', '')
            ma_hp = row.get('Mã học phần', '')
            key = (msv, ma_hp)
            if key not in merged_dict:
                merged_dict[key] = row.copy()
            else:
                existing_row = merged_dict[key]
                for col in final_order:
                    val_old = str(existing_row.get(col, "")).strip()
                    val_new = str(row.get(col, "")).strip()
                    if col == 'Tên học phần':
                        if (not val_old or val_old.lower() == 'result') and (val_new and val_new.lower() != 'result'):
                            existing_row[col] = val_new
                    else:
                        if not val_old and val_new:
                            existing_row[col] = val_new
                            
        results_list = list(merged_dict.values())
        df_results = pl.DataFrame(results_list).select(final_order)
        output_path = os.path.join(data_dir, 'Result.xlsx')
        
        final_df = df_results.to_pandas()
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='DuLieuTrichXuat')
            worksheet = writer.sheets['DuLieuTrichXuat']
            for idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                adjusted_width = min((max_length + 2), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        return output_path
    else:
        raise ValueError("Quá trình quét kết thúc. Không tìm thấy dữ liệu nào trùng khớp với danh sách TKSV.")

def gom_diem_uni_logic(msv_path, data_dir):
    wb_msv = CalamineWorkbook.from_path(msv_path)
    sheet_msv = wb_msv.get_sheet_by_index(0)
    try: raw_msv = sheet_msv.to_python(skip_empty_area=False)
    except TypeError: raw_msv = sheet_msv.to_python()
    
    if not raw_msv or len(raw_msv) < 2:
        raise ValueError("File MSV rỗng.")
        
    raw_ids = []
    for row in raw_msv[1:]:
        if len(row) > 0 and row[0] is not None:
            val = str(row[0]).strip().upper()
            if val.endswith('.0') and val[:-2].lstrip('-').isdigit():
                val = val[:-2]
            if val: raw_ids.append(val)
                
    msv_set = set(raw_ids)
    msv_set.discard("") 
    msv_list = list(msv_set)
    
    if not msv_list:
        raise ValueError("Không tìm thấy MSV hợp lệ trong file MSV.")

    target_files = []
    for root_dir, dirs, files in os.walk(data_dir):
        for file in files:
            file_lower = file.lower()
            if file_lower.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')) and not file.startswith('~$') and file_lower != "result.xlsx":
                target_files.append(os.path.join(root_dir, file))

    if not target_files:
        raise ValueError("Không có file dữ liệu nào để quét.")

    all_matched_data = []

    for file_path in target_files:
        filename = os.path.basename(file_path)
        try:
            wb = CalamineWorkbook.from_path(file_path)
            for sheet_name in wb.sheet_names:
                sheet = wb.get_sheet_by_name(sheet_name)
                try: raw_data = sheet.to_python(skip_empty_area=False)
                except TypeError: raw_data = sheet.to_python()
                
                if not raw_data: continue
                max_cols = max(len(row) for row in raw_data)
                if max_cols == 0: continue
                
                str_data = []
                for row in raw_data:
                    str_row = [str(val).strip() if val is not None else "" for val in row]
                    if len(str_row) < max_cols:
                        str_row.extend([""] * (max_cols - len(str_row)))
                    str_data.append(str_row)
                    
                df = pl.DataFrame(str_data, orient="row")
                df = df.select(pl.all().cast(pl.Utf8))
                
                exprs = [
                    pl.col(c).str.strip_chars().str.to_uppercase().is_in(msv_list)
                    for c in df.columns
                ]
                
                mask = pl.any_horizontal(*exprs)
                matched_df = df.filter(mask)

                if len(matched_df) > 0:
                    matched_rows = matched_df.to_pandas()
                    matched_rows = matched_rows.fillna("")
                    matched_rows.insert(0, 'Nguồn_File', filename)
                    matched_rows.insert(1, 'Nguồn_Sheet', sheet_name)
                    matched_rows.insert(2, 'Đường_Dẫn_File', filename)
                    
                    all_matched_data.append(matched_rows)
        except Exception:
            pass

    if all_matched_data:
        final_df = pd.concat(all_matched_data, ignore_index=True)
        output_path = os.path.join(data_dir, "Result.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='DuLieuTrichXuat')
            worksheet = writer.sheets['DuLieuTrichXuat']
            for idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                adjusted_width = min((max_length + 2), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        return output_path
    else:
        raise ValueError("Quá trình quét kết thúc. Không tìm thấy dữ liệu nào trùng khớp với danh sách MSV.")

def check_dk_dangky_logic(folder_path, has_header):
    output_files = []
    for f in os.listdir(folder_path):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            input_file = os.path.join(folder_path, f)
            df = pd.read_excel(input_file, header=None)
            
            if has_header:
                header_row = df.iloc[0:1].copy()
                df = df.iloc[1:].copy()
            else: header_row = pd.DataFrame()

            while df.shape[1] < 15: df[df.shape[1]] = None
            B, D, H, I, L, M, N, O = 1, 3, 7, 8, 11, 12, 13, 14
            df = df.sort_values(by=[B, D], ascending=[True, True]).reset_index(drop=True)

            def extract_before_last_dot(val):
                if pd.isna(val): return ""
                val_str = str(val)
                if '.' in val_str: return val_str.rsplit('.', 1)[0]
                return val_str

            df[M] = df[D].apply(extract_before_last_dot)

            temp_h_str = df[H].astype(str).str.replace(',', '.', regex=False)
            temp_h = pd.to_numeric(temp_h_str, errors='coerce').astype(float)
            temp_i_str = df[I].astype(str).str.replace(',', '.', regex=False)
            temp_i = pd.to_numeric(temp_i_str, errors='coerce').astype(float)
            temp_l = pd.to_datetime(df[L], errors='coerce', dayfirst=True)

            df[N] = ""
            for (val_b, val_m), group in df.groupby([B, M]):
                if pd.isna(val_b) and pd.isna(val_m): continue
                valid_dates_idx = temp_l[group.index].dropna().index
                if not valid_dates_idx.empty: latest_idx = temp_l[valid_dates_idx].idxmax()
                else: latest_idx = group.index[0]

                h_val = temp_h.at[latest_idx]
                i_val = temp_i.at[latest_idx]

                if pd.notna(h_val) and pd.notna(i_val):
                    try:
                        if float(h_val) >= 40.0 and float(i_val) >= 40.0: df.at[latest_idx, N] = "YES"
                        else: df.at[latest_idx, N] = "NO"
                    except ValueError: df.at[latest_idx, N] = "NO"
                else: df.at[latest_idx, N] = "NO"

            df[O] = ""
            for (val_b, val_m), group in df.groupby([B, M]):
                yes_indices = group[df.loc[group.index, N] == "YES"].index
                if len(yes_indices) > 0:
                    unique_dates = temp_l[group.index].dropna().nunique()
                    for idx in yes_indices:
                        if unique_dates > 1: df.at[idx, O] = "TL"
                        else: df.at[idx, O] = "TL1"

            if has_header:
                while header_row.shape[1] < df.shape[1]: header_row[header_row.shape[1]] = ""
                if pd.isna(header_row.at[0, M]) or header_row.at[0, M] == "": header_row.at[0, M] = "Kết quả M"
                if pd.isna(header_row.at[0, N]) or header_row.at[0, N] == "": header_row.at[0, N] = "Đánh giá N"
                if pd.isna(header_row.at[0, O]) or header_row.at[0, O] == "": header_row.at[0, O] = "Phân loại O"
                df = pd.concat([header_row, df], ignore_index=True)

            name, ext = os.path.splitext(f)
            out_filename = f"{name}_finish{ext}"
            out_filepath = os.path.join(folder_path, out_filename)
            df.to_excel(out_filepath, index=False, header=False)
            output_files.append(out_filepath)
            
    if not output_files: raise ValueError("Không có file Excel nào để xử lý.")
    if len(output_files) == 1: return output_files[0]
    zip_path = os.path.join(folder_path, "KetQua_KiemTra_DK.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for fpath in output_files: zipf.write(fpath, os.path.basename(fpath))
    return zip_path

def scrape_ehou_logic(folder_path, status_placeholder=None):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
    import time

    def fast_paste(driver, element, text_data):
        script = """
            var elm = arguments[0];
            elm.value = arguments[1];
            elm.dispatchEvent(new Event('input', { bubbles: true }));
            elm.dispatchEvent(new Event('change', { bubbles: true }));
        """
        driver.execute_script(script, element, text_data)

    input_file = None
    for f in os.listdir(folder_path):
        if f.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')) and not f.startswith('~$'):
            input_file = os.path.join(folder_path, f)
            break
            
    if not input_file: raise ValueError("Không tìm thấy file Excel nguồn trong thư mục!")

    try:
        xls = pd.ExcelFile(input_file)
    except Exception as e:
        raise ValueError(f"Không thể đọc file Excel. Định dạng không hợp lệ: {e}")

    if 'Login' not in xls.sheet_names or 'Data' not in xls.sheet_names:
        raise ValueError("File Excel phải chứa đúng 2 sheet: 'Login' và 'Data'.")

    df_login = pd.read_excel(xls, sheet_name='Login', header=None)
    if df_login.empty or df_login.shape[1] < 2:
        raise ValueError("Sheet 'Login' trống hoặc thiếu cột Dữ liệu.")
    
    first_cell = str(df_login.iloc[0, 0]).lower()
    if "user" in first_cell or "tài khoản" in first_cell or "tk" in first_cell:
        if df_login.shape[0] < 2:
            raise ValueError("Sheet 'Login' chỉ có tiêu đề mà không có dữ liệu thật.")
        username = str(df_login.iloc[1, 0]).strip().lstrip("'")
        password_raw = str(df_login.iloc[1, 1]).strip().lstrip("'")
    else:
        username = str(df_login.iloc[0, 0]).strip().lstrip("'")
        password_raw = str(df_login.iloc[0, 1]).strip().lstrip("'")
    
    password = password_raw.strip("/") 
    if not username or username == 'nan':
        raise ValueError("Không đọc được Tài khoản trong Sheet 'Login'. Hãy đảm bảo điền Account ở ô A2, Password ô B2.")
    
    df_data = pd.read_excel(xls, sheet_name='Data', dtype=str)
    total_rows = len(df_data)
    batch_size = 100 

    result_path = os.path.join(folder_path, "Ket_Qua_Tra_Cuu.xlsx")
    temp_csv_path = os.path.join(folder_path, "temp_backup_result.csv")

    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        raise RuntimeError(f"Lỗi khởi chạy môi trường duyệt web ẩn. Hãy đảm bảo Server có 'chromium' và 'chromium-driver'.\nChi tiết: {e}")

    wait = WebDriverWait(driver, 25) 

    try:
        if status_placeholder: status_placeholder.info("⚙️ Đang truy cập và đăng nhập hệ thống EHOU...")
        driver.get("https://learning.ehou.edu.vn")
        
        txt_user = wait.until(EC.presence_of_element_located((By.NAME, "username")))
        txt_user.clear()
        txt_user.send_keys(username)

        txt_pass = driver.find_element(By.NAME, "password")
        txt_pass.clear()
        txt_pass.send_keys(password)
        txt_pass.send_keys(Keys.RETURN) 

        time.sleep(5) 
        
        data_processed = False

        for start_idx in range(0, total_rows, batch_size):
            end_idx = min(start_idx + batch_size, total_rows)
            
            chunk = df_data.iloc[start_idx:end_idx] 
            chunk_accounts = []
            chunk_courses = []
            
            for index, row in chunk.iterrows():
                account_val = str(row.iloc[0]).strip()
                if not account_val or account_val == 'nan': continue
                    
                for col_idx in range(1, 16):
                    if col_idx >= len(row): break
                    course_val = str(row.iloc[col_idx]).strip()
                    if course_val and course_val != 'nan':
                        chunk_accounts.append(account_val)
                        chunk_courses.append(course_val)
                        
            if not chunk_accounts: continue
            data_processed = True
            if status_placeholder: status_placeholder.info(f"🔎 Đang tra cứu dữ liệu: Lô từ dòng {start_idx + 1} đến {end_idx}...")
                
            account_str_to_paste = " ".join(chunk_accounts)
            course_str_to_paste = " ".join(chunk_courses)
            
            driver.get("https://learning.ehou.edu.vn/grade/report/overview/advance.php")
            
            try:
                txt_account = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), 'Cột tài khoản')]/following-sibling::*[self::input or self::textarea]")))
                fast_paste(driver, txt_account, account_str_to_paste)

                txt_course = driver.find_element(By.XPATH, "//label[contains(text(), 'Cột mã môn')]/following-sibling::*[self::input or self::textarea]")
                fast_paste(driver, txt_course, course_str_to_paste)

                driver.execute_script("""
                    var oldTables = document.getElementsByTagName('table');
                    for (var i = oldTables.length - 1; i >= 0; i--) {
                        oldTables[i].parentNode.removeChild(oldTables[i]);
                    }
                """)

                txt_course.send_keys(Keys.RETURN)
                
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, "//table[contains(., 'Mã lớp môn')]")))
                    time.sleep(0.5) 
                except: pass 

                js_extract_script = """
                    var tables = document.getElementsByTagName('table');
                    var targetTable = null;
                    for (var k = 0; k < tables.length; k++) {
                        if (tables[k].innerText.includes('Mã lớp môn') && tables[k].innerText.includes('Học viên')) {
                            targetTable = tables[k];
                            break;
                        }
                    }
                    if (!targetTable) return null;
                    var data = [];
                    var rows = targetTable.rows;
                    for (var i = 0; i < rows.length; i++) {
                        var rowData = [];
                        var cols = rows[i].cells;
                        for (var j = 0; j < cols.length; j++) {
                            rowData.push(cols[j].innerText.trim().replace(/\\n/g, ' '));
                        }
                        data.push(rowData);
                    }
                    return data;
                """
                
                table_data = driver.execute_script(js_extract_script)
                df_result = None

                if table_data and len(table_data) > 1:
                    max_cols = max(len(r) for r in table_data)
                    cleaned_data = [r + [""] * (max_cols - len(r)) for r in table_data]
                    
                    headers = cleaned_data[0]
                    final_headers = []
                    for i, h in enumerate(headers):
                        if not h: h = f"Cot_trong_{i}"
                        if h in final_headers: h = f"{h}_{i}"
                        final_headers.append(h)
                        
                    temp_df = pd.DataFrame(cleaned_data[1:], columns=final_headers)
                    expected_columns = [
                        "TT", "Học viên", "Họ tên sinh viên", "Mã lớp môn", "Tên lớp môn", 
                        "Lớp quản lý", "Email", "Quá trình", "Kiểm tra giữa kỳ", 
                        "Kết thúc học phần", "Tổng kết", "Thời gian bắt đầu"
                    ]
                    
                    if "STT" in temp_df.columns and "TT" not in temp_df.columns:
                        temp_df = temp_df.rename(columns={"STT": "TT"})
                        
                    extracted_cols = []
                    for expected in expected_columns:
                        if expected in temp_df.columns: extracted_cols.append(expected)
                        else:
                            for c in temp_df.columns:
                                if expected.lower() in str(c).lower() or str(c).lower() in expected.lower():
                                    if c not in extracted_cols:
                                        temp_df = temp_df.rename(columns={c: expected})
                                        extracted_cols.append(expected)
                                    break
                    
                    if extracted_cols: df_result = temp_df[extracted_cols]
                    else: df_result = temp_df 

                if df_result is None or df_result.empty:
                    body_text = driver.find_element(By.TAG_NAME, "body").text
                    if "Apereo" in body_text and "Username" in body_text:
                        raise Exception("BỊ VĂNG KHỎI TÀI KHOẢN (Session Timeout)")
                    
                    df_result = pd.DataFrame([{
                        "Học viên": f"Lô dòng {start_idx + 1} - {end_idx}", 
                        "Mã lớp môn": "Nhiều môn", 
                        "Ghi chú": "Không có kết quả / Web lỗi / Không tìm thấy"
                    }])

                if not os.path.exists(temp_csv_path):
                    df_result.to_csv(temp_csv_path, index=False, mode='w', encoding='utf-8-sig')
                else:
                    df_result.to_csv(temp_csv_path, index=False, mode='a', header=False, encoding='utf-8-sig')

            except Exception as e:
                df_err = pd.DataFrame([{
                    "Học viên": f"Lô dòng {start_idx + 1} - {end_idx}", 
                    "Mã lớp môn": "Nhiều môn", 
                    "Ghi chú": f"LỖI HỆ THỐNG: {str(e)}"
                }])
                if not os.path.exists(temp_csv_path): df_err.to_csv(temp_csv_path, index=False, mode='w', encoding='utf-8-sig')
                else: df_err.to_csv(temp_csv_path, index=False, mode='a', header=False, encoding='utf-8-sig')
        
        if not data_processed:
            raise ValueError("File Excel có Sheet Data nhưng dữ liệu trống hoặc sai chuẩn (thiếu Tài khoản ở Cột đầu tiên).")

        if os.path.exists(temp_csv_path):
            final_df = pd.read_csv(temp_csv_path)
            final_df.to_excel(result_path, sheet_name="Ket_Qua_Tong_Hop", index=False, engine='openpyxl')
            os.remove(temp_csv_path) 
            return result_path
        else:
            raise ValueError("Quá trình quét kết thúc nhưng không có dữ liệu để xuất file.")

    finally:
        if driver: driver.quit()

def check_khlm_logic(folder_path):
    file_path = os.path.join(folder_path, "Data_SLLM.xlsx")
    if not os.path.exists(file_path): raise ValueError("Không tìm thấy tệp Data_SLLM.xlsx!")
    output_path = os.path.join(folder_path, "Data_SLLM_Finish.xlsx")
    try: df_data = pd.read_excel(file_path, sheet_name='Data', usecols="D,H")
    except ValueError: raise Exception("Không tìm thấy Sheet 'Data' hoặc thiếu cột D, H.")
    df_data.columns = ['Lop', 'MaMon']
    df_data = df_data.dropna(subset=['MaMon'])
    df_data['Lop'] = df_data['Lop'].astype(str).str.strip()
    df_data['MaMon'] = df_data['MaMon'].astype(str).str.strip()
    grouped_data = df_data.groupby('MaMon')['Lop'].apply(lambda x: ', '.join(sorted(set(x)))).reset_index()
    list_b_mamon_data = grouped_data['MaMon'].tolist()
    list_a_lop_data = grouped_data['Lop'].tolist()
    try: df_tk = pd.read_excel(file_path, sheet_name='ThongKe', usecols="A,B")
    except ValueError: raise Exception("Không tìm thấy Sheet 'ThongKe' hoặc thiếu cột A, B.")
    df_tk.columns = ['TenLop', 'MaMon']
    df_tk = df_tk.dropna(subset=['MaMon'])
    df_tk['TenLop'] = df_tk['TenLop'].astype(str).str.strip()
    df_tk['MaMon'] = df_tk['MaMon'].astype(str).str.strip()
    grouped_tk = df_tk.groupby('MaMon')['TenLop'].apply(lambda x: ', '.join(sorted(set(x)))).reset_index()
    list_e_mamon_tk = grouped_tk['MaMon'].tolist()
    list_d_lop_tk = grouped_tk['TenLop'].tolist()
    list_g_lop_thieu, list_h_mamon_thieu, list_i_mamon_tuong_ung = [], [], []
    dict_tk = dict(zip(list_e_mamon_tk, list_d_lop_tk))
    for i, ma_mon_data in enumerate(list_b_mamon_data):
        classes_in_data = [c.strip() for c in list_a_lop_data[i].split(',')]
        if ma_mon_data not in dict_tk:
            list_h_mamon_thieu.append(ma_mon_data)
            for cls in classes_in_data:
                list_g_lop_thieu.append(cls)
                list_i_mamon_tuong_ung.append(ma_mon_data)
        else:
            classes_in_tk = [c.strip() for c in dict_tk[ma_mon_data].split(',')]
            for cls in classes_in_data:
                if cls not in classes_in_tk:
                    list_g_lop_thieu.append(cls)
                    list_i_mamon_tuong_ung.append(ma_mon_data)
    result_df = pd.DataFrame({
        'Cột A (Lớp_D)': pd.Series(list_a_lop_data), 'Cột B (Mã môn_D)': pd.Series(list_b_mamon_data), 'Cột C (Trống)': pd.Series([], dtype=str),
        'Cột D (Lớp_TK)': pd.Series(list_d_lop_tk), 'Cột E (Mã môn_TK)': pd.Series(list_e_mamon_tk), 'Cột F (Trống)': pd.Series([], dtype=str),
        'Cột G (Lớp_D thiếu trong Lớp_TK)': pd.Series(list_g_lop_thieu), 'Cột H (Mã môn_D thiếu trong Mã môn_TK)': pd.Series(list_h_mamon_thieu),
        'Cột I (Mã môn tương ứng với cột G)': pd.Series(list_i_mamon_tuong_ung),
    }).fillna("")
    shutil.copy(file_path, output_path)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        result_df.to_excel(writer, sheet_name='Result', index=False)
        worksheet = writer.sheets['Result']
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            worksheet.column_dimensions[column].width = min(max_length + 2, 40)
    wb_out = load_workbook(output_path)
    apply_full_border(wb_out['Result'])
    wb_out.save(output_path)
    return output_path

def fill_khlm_logic(folder_path, keywords_str):
    target_file = ""
    for f in os.listdir(folder_path):
        if f.endswith(".xlsx") and not f.startswith("~$"):
            try:
                temp_wb = load_workbook(os.path.join(folder_path, f), read_only=True)
                lower_sheets = [s.lower() for s in temp_wb.sheetnames]
                if 'khlm' in lower_sheets and 'data' in lower_sheets:
                    target_file = os.path.join(folder_path, f)
                    break
            except: continue
    if not target_file: raise ValueError("Không tìm thấy file Excel nào chứa đủ 2 sheet 'KHLM' và 'data'!")
    processor = ExcelDataProcessor(input_file=target_file, keywords_str=keywords_str, log_callback=lambda msg: None)
    success = processor.run_all_steps()
    if not success: raise ValueError("Xử lý thất bại. Vui lòng kiểm tra lại định dạng file.")
    zip_path = os.path.join(folder_path, "KetQua_KHLM_TongHop.zip")
    output_files = ["Ketqua_KHLM.xlsx", "Ketquahoclai_KHLM.xlsx", "Data_Merge.xlsx", "LopMon_Ketqua.xlsx", "Data_fill_Finish.xlsx"]
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for fname in output_files:
            fpath = os.path.join(folder_path, fname)
            if os.path.exists(fpath): zipf.write(fpath, fname)
    return zip_path

def extract_courses_logic(folder_path):
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$') and f != "Courses_List.xlsx"]
    if not files: raise ValueError("Không có file Excel nào để xử lý!")
    output_file = os.path.join(folder_path, "Courses_List.xlsx")
    wb_out = Workbook(); wb_out.remove(wb_out.active)
    for idx, file_name in enumerate(files):
        try: parts = file_name.split('- '); major_name = parts[1].split('.')[0].strip() if len(parts) > 1 else file_name.split('.')[0].strip()
        except: major_name = file_name[:25] 
        ws_major = wb_out.create_sheet(title=major_name[:30])
        unique_courses = set()
        wb_src = load_workbook(os.path.join(folder_path, file_name), data_only=True)
        for sheet_name in wb_src.sheetnames:
            ws_src = wb_src[sheet_name]
            for r in range(16, ws_src.max_row + 1):
                val_a = str(ws_src.cell(row=r, column=1).value or "").strip().lower()
                if "lưu ý:" in val_a: break
                col_b, col_c, col_d = ws_src.cell(row=r, column=2).value, ws_src.cell(row=r, column=3).value, ws_src.cell(row=r, column=4).value 
                if col_b and col_c: unique_courses.add(f"{str(col_b).strip()}_{str(col_c).strip()}_{str(col_d or 0).strip()} Tín chỉ")
        sorted_list = sorted(list(unique_courses))
        for i, course in enumerate(sorted_list, 1): ws_major.cell(row=i, column=1, value=course)
        apply_full_border(ws_major); auto_fit_columns(ws_major)
    wb_out.save(output_file)
    return output_file

def fill_sllm_logic(folder_path):
    file_path = os.path.join(folder_path, "Data_SLLM.xlsx")
    output_file_path = os.path.join(folder_path, "Data_SLLM_Finish.xlsx")
    if not os.path.exists(file_path): 
        raise ValueError("Không tìm thấy tệp Data_SLLM.xlsx!")
        
    wb = load_workbook(file_path)
    if "Data" not in wb.sheetnames or "ThongKe" not in wb.sheetnames: 
        raise ValueError("Tệp Data_SLLM.xlsx phải chứa sheet 'Data' và 'ThongKe'!")
        
    ws_data = wb["Data"]
    data_map = {}
    unique_regions = set()  # Tập hợp lưu trữ các mã duy nhất quét được từ cột L
    
    # --- BƯỚC 1: QUÉT MÃ TẠI CỘT L VÀ GOM NHÓM DỮ LIỆU TRÊN SHEET DATA ---
    for r in range(2, ws_data.max_row + 1):
        # Đọc trực tiếp mã tại cột L (Cột 12), loại bỏ khoảng trắng thừa
        cell_l = ws_data.cell(row=r, column=12).value
        region = str(cell_l or "").strip()
        
        # Nếu dòng này có mã (không bị trống), thêm vào tập hợp các mã duy nhất
        if region:
            unique_regions.add(region)
            
        ma_lm = str(ws_data.cell(row=r, column=4).value or "").strip().lower()
        ma_hp = str(ws_data.cell(row=r, column=8).value or "").strip().lower()
        
        if ma_lm and ma_hp:
            key = (ma_hp, ma_lm)
            if key not in data_map: 
                data_map[key] = {}
            
            # Tích lũy số lượng cho mã cụ thể tại cột L của cặp (mã học phần, mã lớp môn)
            if region:
                data_map[key][region] = data_map[key].get(region, 0) + 1

    # Sắp xếp danh sách mã theo thứ tự chữ cái để tiêu đề hiển thị ngăn nắp (VD: DNP, HCM,...)
    regions_list = sorted(list(unique_regions))
    
    # --- BƯỚC 2: CẬP NHẬT TIÊU ĐỀ ĐỘNG TRÊN SHEET THONGKE ---
    ws_tk = wb["ThongKe"]
    
    # Điền các mã duy nhất vào dòng 1, bắt đầu từ cột C (Cột 3) trở đi
    for idx, region_code in enumerate(regions_list):
        ws_tk.cell(row=1, column=3 + idx, value=region_code)
        
    # --- BƯỚC 3: TỔNG HỢP VÀ THỐNG KÊ DỮ LIỆU ĐỘNG SANG SHEET THONGKE ---
    for r in range(2, ws_tk.max_row + 1):
        ma_hp_tk = str(ws_tk.cell(row=r, column=2).value or "").strip().lower()
        ma_lm_str = str(ws_tk.cell(row=r, column=1).value or "").strip()
        
        # Khởi tạo từ điển lưu tổng số lượng của từng mã cho dòng thống kê hiện tại
        sums_by_region = {region: 0 for region in regions_list}
        
        if ma_hp_tk and ma_lm_str:
            codes = [c.strip().lower() for c in ma_lm_str.split(',')]
            for code in codes:
                key = (ma_hp_tk, code)
                if key in data_map:
                    # Cộng dồn số lượng dựa trên các mã có trong danh sách động
                    for region in regions_list:
                        sums_by_region[region] += data_map[key].get(region, 0)
                        
        # Ghi các giá trị tổng hợp được vào các cột tương ứng trên sheet ThongKe
        for idx, region in enumerate(regions_list):
            ws_tk.cell(row=r, column=3 + idx, value=sums_by_region[region])
            
    # --- BƯỚC 4: ĐỊNH DẠNG VÀ LƯU FILE ---
    for sheet in [ws_data, ws_tk]: 
        apply_full_border(sheet)
        auto_fit_columns(sheet)
        
    wb.save(output_file_path)
    return output_file_path

def compare_data_logic(folder_path):
    file_xlsb, file_src, output_file = os.path.join(folder_path, "Data.xlsb"), os.path.join(folder_path, "Data_Source.xlsx"), os.path.join(folder_path, "Compared_Result.xlsx")
    if not os.path.exists(file_xlsb) or not os.path.exists(file_src): raise ValueError("Không tìm thấy tệp Data.xlsb hoặc Data_Source.xlsx!")
    wb_src = load_workbook(file_src, data_only=True)
    if "DSSV" not in wb_src.sheetnames: raise ValueError("Thiếu sheet DSSV trong Data_Source.xlsx")
    ws_src = wb_src["DSSV"]
    source_codes = {str(ws_src.cell(row=r, column=10).value).strip().lower() for r in range(2, ws_src.max_row + 1) if ws_src.cell(row=r, column=10).value}
    wb_out = Workbook(); ws_out = wb_out.active; ws_out.title = "Result"; headers_added = False
    with open_xlsb(file_xlsb) as wb_bin:
        with wb_bin.get_sheet(1) as sheet:
            for row in sheet.rows():
                val_a = str(row[0].v).strip().lower() if row[0].v is not None else ""
                if not headers_added:
                    ws_out.append([c.v for c in row]); headers_added = True; continue
                if val_a in source_codes: ws_out.append([c.v for c in row])
    apply_full_border(ws_out); auto_fit_columns(ws_out); wb_out.save(output_file)
    return output_file

def filter_sv_logic(folder_path):
    f_dk, f_src = os.path.join(folder_path, "DanhSachDangKy.xlsx"), os.path.join(folder_path, "Data_Source.xlsx")
    if not os.path.exists(f_dk) or not os.path.exists(f_src): raise ValueError("Thiếu file DanhSachDangKy.xlsx hoặc Data_Source.xlsx")
    wb_src = load_workbook(f_src, data_only=True); ws_src = wb_src["DSSV"]
    sv_set = {str(ws_src.cell(row=r, column=4).value).strip().lower() for r in range(2, ws_src.max_row + 1) if ws_src.cell(row=r, column=4).value}
    wb_dk = load_workbook(f_dk, data_only=True)
    reg_data, learn_keys, h_reg, h_learn = [], set(), [], []
    if "Dangky" in wb_dk.sheetnames:
        ws = wb_dk["Dangky"]; h_reg = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[1] and str(r[1]).strip().lower() in sv_set: reg_data.append(r)
    if "Danghoc" in wb_dk.sheetnames:
        ws = wb_dk["Danghoc"]; h_learn = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[1] and str(r[1]).strip().lower() in sv_set: learn_keys.add((str(r[1] or "").strip().lower(), str(r[6] or "").strip().lower(), str(r[14] or "").strip().lower()))
    m_rows, n_rows = [], []
    for r in reg_data:
        rk = (str(r[1] or "").strip().lower(), str(r[2] or "").strip().lower(), str(r[19] or "").strip().lower())
        if rk in learn_keys: m_rows.append(r)
        else: n_rows.append(r)
    output_file = os.path.join(folder_path, "Filter_Result.xlsx"); wb_res = Workbook()
    configs = [("Register_List", reg_data, h_reg), ("matching", m_rows, h_reg), ("not_matching", n_rows, h_reg)]
    for i, (t, d, h) in enumerate(configs):
        ws = wb_res.active if i == 0 else wb_res.create_sheet(t); ws.title = t; ws.append(h)
        for row in d: ws.append(row)
        apply_full_border(ws); auto_fit_columns(ws)
    wb_res.save(output_file)
    return output_file

def export_khhtct_logic(folder_path):
    tf = os.path.join(folder_path, "Merged_GK300.xlsx")
    if not os.path.exists(tf): raise ValueError("Cần file Merged_GK300.xlsx")
    wb = load_workbook(tf); ws_d = wb["DSSV_GK300"]; ws_k = wb["KHHT_GK300"]; kd = {}
    for r in range(2, ws_k.max_row + 1):
        k, m = ws_k.cell(row=r, column=1).value, ws_k.cell(row=r, column=4).value 
        if k and m:
            sk = str(k).strip().lower()
            if sk not in kd: kd[sk] = []
            kd[sk].append(m)
    if "KHHTCT_GK300" in wb.sheetnames: del wb["KHHTCT_GK300"]
    ws_n = wb.create_sheet("KHHTCT_GK300")
    hd = ["STT", "Lớp", "Tài khoản SV", "Họ và đệm", "Tên", "Ngày sinh", "Mã sinh viên", "Mã lớp môn"]
    for c, h in enumerate(hd, 1): ws_n.cell(row=1, column=c, value=h).font = Font(bold=True)
    cr, stt = 2, 1
    for r in range(2, ws_d.max_row + 1):
        l = ws_d.cell(row=r, column=17).value
        if l:
            sl = str(l).strip()
            if len(sl) >= 4:
                mk = (sl[0] + sl[-3:]).lower()
                if mk in kd:
                    for ml in kd[mk]:
                        dr = [stt, sl, ws_d.cell(row=r, column=4).value, ws_d.cell(row=r, column=12).value, ws_d.cell(row=r, column=13).value, ws_d.cell(row=r, column=14).value, ws_d.cell(row=r, column=10).value, ml]
                        for c, v in enumerate(dr, 1): ws_n.cell(row=cr, column=c, value=v)
                        cr += 1; stt += 1
    apply_full_border(ws_n); auto_fit_columns(ws_n); wb.save(tf)
    return tf

def export_dssv_logic(folder_path):
    sf, df = os.path.join(folder_path, "Data_Source.xlsx"), os.path.join(folder_path, "Merged_GK300.xlsx")
    if not os.path.exists(sf) or not os.path.exists(df): raise ValueError("Cần file Data_Source.xlsx và Merged_GK300.xlsx")
    wb_d = load_workbook(df); ws_k = wb_d["KHHT_GK300"]; cs = {str(ws_k.cell(row=r, column=12).value).strip().lower() for r in range(2, ws_k.max_row + 1) if ws_k.cell(row=r, column=12).value} 
    wb_s = load_workbook(sf, data_only=True); ws_s = wb_s["DSSV"]
    if "DSSV_GK300" in wb_d.sheetnames: del wb_d["DSSV_GK300"]
    ws_n = wb_d.create_sheet("DSSV_GK300")
    for c in range(1, ws_s.max_column + 1): ws_n.cell(row=1, column=c, value=ws_s.cell(row=1, column=c).value).font = Font(bold=True)
    cr = 2
    for r in range(2, ws_s.max_row + 1):
        vu = ws_s.cell(row=r, column=17).value
        if vu:
            su = str(vu).strip()
            if len(su) >= 4:
                mk = (su[0] + su[-3:]).lower()
                if mk in cs:
                    for c in range(1, ws_s.max_column + 1): ws_n.cell(row=cr, column=c, value=ws_s.cell(row=r, column=c).value)
                    cr += 1
    apply_full_border(ws_n); auto_fit_columns(ws_n); wb_d.save(df)
    return df

def export_khht_logic(folder_path):
    fs = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$') and f not in ["Merged_GK300.xlsx", "Data_Source.xlsx", "ClassName.xlsx", "Filter_Result.xlsx", "DanhSachDangKy.xlsx"]]
    if not fs: raise ValueError("Không có file dữ liệu GK300!")
    op = os.path.join(folder_path, "Merged_GK300.xlsx"); wb = Workbook(); ws = wb.active; ws.title = "KHHT_GK300"
    hd = ["Mã LT", "HK", "Tên HP", "Mã HP", "Số TC", "Ngày bắt đầu (Dự kiến)", "Ngày kết thúc (Dự kiến)", "Mã ngành", "Khóa", "Ghi chú", "", "Mã LT (Duy nhất)"]
    for c, h in enumerate(hd, 1): ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    mlt = []
    for f in fs:
        swb = load_workbook(os.path.join(folder_path, f), data_only=True)
        for sn in swb.sheetnames:
            s = swb[sn]; mlt.append(sn); hk, bd, kt = format_hk(s["A6"].value), format_excel_date(s["J11"]), format_excel_date(s["X11"])
            for r in range(13, s.max_row + 1):
                t, m, tc = s.cell(row=r, column=2).value, s.cell(row=r, column=3).value, s.cell(row=r, column=4).value
                if t and m: ws.append([sn, hk, t, m, tc, bd, kt, str(sn)[0], str(sn)[-2:], ""])
    unq = sorted(list(set(mlt)))
    for i, v in enumerate(unq, 2): ws.cell(row=i, column=12, value=v)
    apply_full_border(ws); auto_fit_columns(ws); wb.save(op)
    return op

def process_files_logic(folder_path):
    fs = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$') and f not in ["File_Merged.xlsx", "ClassName.xlsx", "Merged_GK300.xlsx", "Data_Source.xlsx", "Filter_Result.xlsx", "DanhSachDangKy.xlsx"]]
    if not fs: raise ValueError("Không có file để gộp!")
    op = os.path.join(folder_path, "File_Merged.xlsx"); mwb = Workbook(); dws = mwb.active; dws.title = "Sheet_Merged"
    yl = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"); cdr, mcw = 1, {}
    for f in fs:
        swb = load_workbook(os.path.join(folder_path, f), data_only=False)
        for sn in swb.sheetnames:
            sws = swb[sn]
            for cd in sws.column_dimensions.values():
                for ci in range(cd.min, cd.max + 1):
                    if ci <= 36: l = get_column_letter(ci); mcw[l] = max(mcw.get(l, 0), cd.width or 0)
            hrm = {}
            for hr in [8,9,10,11,12]:
                hrm[hr] = cdr
                if sws.row_dimensions[hr].height: dws.row_dimensions[cdr].height = sws.row_dimensions[hr].height
                for c in range(1, 37):
                    sc, dc = sws.cell(row=hr, column=c), dws.cell(row=cdr, column=c, value=sws.cell(row=hr, column=c).value)
                    if sc.has_style: dc.font, dc.border, dc.fill, dc.number_format, dc.alignment = copy(sc.font), copy(sc.border), copy(sc.fill), copy(sc.number_format), copy(sc.alignment)
                    if hr == 8: dc.fill = yl
                cdr += 1
            vdr = [r for r in range(13, sws.max_row + 1) if sws.cell(row=r, column=10).value and "-" in str(sws.cell(row=r, column=10).value).lower() and not sws.row_dimensions[r].hidden]
            if not vdr: continue
            brm = {}
            for sr in vdr:
                brm[sr] = cdr
                if sws.row_dimensions[sr].height: dws.row_dimensions[cdr].height = sws.row_dimensions[sr].height
                for c in range(1, 37):
                    sc, dc = sws.cell(row=sr, column=c), dws.cell(row=cdr, column=c, value=sws.cell(row=sr, column=c).value)
                    if sc.has_style: dc.font, dc.border, dc.fill, dc.number_format, dc.alignment = copy(sc.font), copy(sc.border), copy(sc.fill), copy(sc.number_format), copy(sc.alignment)
                dws.cell(row=cdr, column=37, value=f"{f} -> {sn}"); cdr += 1
            for mr in sws.merged_cells.ranges:
                rir = [r for r in vdr if mr.min_row <= r <= mr.max_row]
                if len(rir) >= 1:
                    mnr, mxr = brm[min(rir)], brm[max(rir)]; mnc, mxc = max(mr.min_col, 1), min(mr.max_col, 36)
                    if mnc <= mxc and (mxr > mnr or mxc > mnc):
                        try: dws.merge_cells(start_row=mnr, start_column=mnc, end_row=mxr, end_column=mxc)
                        except: pass
    for l, w in mcw.items(): dws.column_dimensions[l].width = w
    dws.column_dimensions[get_column_letter(37)].width = 30
    apply_full_border(dws); mwb.save(op)
    return op

def extract_class_names_logic(folder_path):
    fs = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$') and f not in ["Merged_GK300.xlsx", "ClassName.xlsx", "Filter_Result.xlsx"]]
    if not fs: raise ValueError("Không có file để xử lý!")
    op = os.path.join(folder_path, "ClassName.xlsx"); wb = Workbook(); ws = wb.active; cr = 1
    for f in fs:
        swb = load_workbook(os.path.join(folder_path, f), read_only=True)
        for sn in swb.sheetnames: ws.cell(row=cr, column=1, value=sn); cr += 1
    apply_full_border(ws); auto_fit_columns(ws); wb.save(op)
    return op

# ==========================================
# MENU SIDEBAR 
# ==========================================
with st.sidebar:
    st.markdown("<h3 style='color: #1E3A8A; font-weight: 700; margin-top: -15px;'>DATA WORKSPACE</h3>", unsafe_allow_html=True)
    st.markdown("---")
    
    menu_options = {
        "Gộp File Nguồn": "Gộp File Nguồn",
        "Tra cứu điểm sinh viên (Cả lớp 1*)": "Tra cứu điểm sinh viên (Cả lớp)", 
        "Kiểm tra ĐK đăng ký (Cả lớp 2*)": "Kiểm tra điều kiện đăng ký môn học (Cả lớp)", 
        "Kiểm tra tiên quyết": "Kiểm tra tiên quyết", 
        "Gom điểm UNI": "Gom điểm UNI",
        "Gom điểm": "Gom điểm", 
        "Điền KHLM (Updated) (*)": "Điền KHLM (Updated) (*)",
        "Kiểm tra KHLM (*)": "Kiểm tra KHLM", 
        "Lọc KQHT Sinh viên": "Lọc kết quả học tập của sinh viên",
        "Lọc SV Học lại/Cải thiện": "Lọc sinh viên học lại & học cải thiện",
        "Xuất Mã lớp (GK300)": "Xuất Mã lớp theo GK300",
        "Xuất KHHT (1*)": "Xuất KHHT theo GK300 (1*)",
        "Xuất DSSV (2*)": "Xuất DSSV theo GK300 (2*)",
        "Xuất KHHTCT (3*)": "Xuất KHHTCT theo GK300 (3*)",
        "Thống kê Lớp & Môn (*)": "Thống kê số lượng theo lớp/nhóm lớp & môn (*)",
        "Xuất Môn theo Ngành (*)": "Xuất môn theo ngành học (*)"
    }
    
    selected_label = st.radio("CHỌN CHỨC NĂNG BÊN DƯỚI", list(menu_options.keys()))
    choice = menu_options[selected_label]
    
    st.markdown("---")
    st.caption("Ver 11.0 | Stable Calamine Edition")

# ==========================================
# GIAO DIỆN CHÍNH (SINGLE-SCREEN GRID LAYOUT)
# ==========================================
st.markdown(f"""
    <div class="app-header">
        <h2>{choice}</h2>
        <p>Thực thi tự động tác vụ Excel một cách nhanh chóng và chính xác.</p>
    </div>
""", unsafe_allow_html=True)

TEMPLATE_BASE_URL = "https://raw.githubusercontent.com/dangvannghia204/hou-tools/main/templates"

def tpl_link(filename):
    return f"<br><a href='{TEMPLATE_BASE_URL}/{filename}' download='{filename}' style='display: inline-block; margin-top: 10px; margin-right: 10px; padding: 6px 12px; background-color: #EFF6FF; color: #1D4ED8; border: 1px solid #BFDBFE; border-radius: 4px; text-decoration: none; font-weight: bold; font-size: 0.85rem;'>📥 File Mẫu: {filename}</a>"

instructions = {
    "Gộp File Nguồn": f"Đầu vào là file <b>GK300</b> của 1 hoặc nhiều khóa (Mỗi sheet chứa bảng đăng ký).",
    "Gom điểm UNI": f"<b>Trích xuất dữ liệu đa File (Powered by Polars & Rust):</b><br><b>1. File MSV:</b> Tải lên file Excel chứa danh sách Mã SV (ở cột đầu tiên) vào ô bên trái.<br><b>2. Dữ liệu Nguồn:</b> Kéo thả TẤT CẢ các file Excel cần quét vào khu vực bên phải.<br><b>Kết quả:</b> Hệ thống quét vét cạn và trả về file tổng hợp <code>Result.xlsx</code>.{tpl_link('TKSV_Template.xlsx')}",
    "Gom điểm": f"<b>Trích xuất dữ liệu chuẩn Form BGD & Thường (Powered by Polars & Rust):</b><br><b>1. File MSV:</b> Tải lên file chứa danh sách TKSV ở cột 'TKSV'.<br><b>2. Dữ liệu Nguồn:</b> Kéo thả TẤT CẢ các file Excel cần quét vét cạn.<br><b>Kết quả:</b> Hệ thống tự động lấy điểm, gộp dòng, và xuất file tổng hợp.{tpl_link('TKSV_Template.xlsx')}",
    "Tra cứu điểm sinh viên (Cả lớp)": f"<b>Tự động Scraping EHOU (Headless):</b><br><br><b>File yêu cầu:</b> Excel chuẩn bị sẵn với 2 sheet:<br><b>1. Sheet 'Login':</b> Ô A2 (Tài khoản), Ô B2 (Mật khẩu).<br><b>2. Sheet 'Data':</b> Cột 1 (Tài khoản SV), Các cột sau chứa mã môn học.{tpl_link('TraCuuDiem_Template.xlsx')}",
    "Kiểm tra điều kiện đăng ký môn học (Cả lớp)": f"<b>Yêu cầu file:</b> Sử dụng kết quả: Tra cứu điểm sinh viên (Cả lớp 1*).<br><b>Tùy chọn:</b> Có thể xác định file có chứa dòng tiêu đề hay không.<br><b>Kết quả:</b> Đánh giá điều kiện đạt (YES/NO) và phân loại trạng thái (TL/TL1).",
    "Kiểm tra tiên quyết": f"<b>Yêu cầu file:</b> Upload đồng thời 2 file <code>Ketqua.xlsx</code> và <code>dktq.xlsx</code>.<br><br><b>Kết quả:</b> Đối soát môn tiên quyết và trả về file nén chứa <code>Ketqua_Finish.xlsx</code> và <code>dktq_Finish.xlsx</code>.{tpl_link('Ketqua_Template.xlsx')}{tpl_link('dktq_Template.xlsx')}",
    "Điền KHLM (Updated) (*)": f"<b>Yêu cầu file:</b> <code>Data_fill.xlsx</code><br><br><b>Sheet KHLM:</b> Cần có các cột <code>TenLop</code>, <code>MaMon</code>, <code>DiaPhuongKHL</code>, <code>DiaPhuongHL</code>.<br><b>Sheet Data:</b> Cần có các cột <code>LopLT</code>, <code>MaMon</code>, <code>MaTram</code>, <code>MSV</code>.{tpl_link('Data_fill_Template.xlsx')}",
    "Kiểm tra KHLM": f"<b>Yêu cầu file:</b> <code>Data_SLLM.xlsx</code><br><br><b>Sheet Data:</b> Cột D (Lớp), Cột H (Mã môn).<br><b>Sheet ThongKe:</b> Cột A (Lớp), Cột B (Mã môn).<br><br><b>Kết quả:</b> Đối soát lộ trình thiếu kèm với tên môn tương ứng.{tpl_link('Data_SLLM_Template.xlsx')}",
    "Lọc kết quả học tập của sinh viên": f"<b>Yêu cầu:</b> <code>Data_Source.xlsx</code> & <code>Data.xlsb</code><br><br><b>Data_Source.xlsx:</b> Sheet 'DSSV', Cột Q (Lớp), Cột J (Mã SV), Cột D (TK SV).<br><b>Data.xlsb:</b> File nhị phân, Cột A (Mã SV).{tpl_link('Data_Source_Template.xlsx')}{tpl_link('Data_Template.xlsb')}",
    "Lọc sinh viên học lại & học cải thiện": f"<b>Yêu cầu:</b> <code>Data_Source.xlsx</code> & <code>DanhSachDangKy.xlsx</code><br><br><b>DanhSachDangKy.xlsx:</b> Sheet 'Dangky: B,C,T' và 'Danghoc: B,G,O', Cột B (TK SV), Cột C&G (Mã môn), Cột T&O (Số TC).{tpl_link('Data_Source_Template.xlsx')}{tpl_link('DanhSachDangKy_Template.xlsx')}",
    "Xuất Mã lớp theo GK300": f"Đầu vào là file <b>GK300</b> của 1 hoặc nhiều khóa.",
    "Xuất KHHT theo GK300 (1*)": f"Đầu vào là file <b>GK300</b> của 1 hoặc nhiều khóa.",
    "Xuất DSSV theo GK300 (2*)": f"<b>Yêu cầu:</b> <code>Merged_GK300.xlsx</code> tạo từ Xuất KHHT (1*) & <code>Data_Source.xlsx</code><br><b>Merged_GK300.xlsx:</b> Sheet 'KHHT_GK300', Cột L(12) là Mã LT.{tpl_link('Data_Source_Template.xlsx')}",
    "Xuất KHHTCT theo GK300 (3*)": f"<b>Yêu cầu:</b> Sử dụng kết quả Xuất DSSV (2*), File: <code>Merged_GK300.xlsx </code>.",
    "Thống kê số lượng theo lớp/nhóm lớp & môn (*)": f"<b>Yêu cầu:</b> File <code>Data_SLLM.xlsx</code><br><b>Sheet Data:</b> Cột L (Mã trạm).<br><b>Sheet ThongKe:</b> Cột A (Tên lớp), B (Mã môn).{tpl_link('Data_SLLM_Template.xlsx')}",
    "Xuất môn theo ngành học (*)": f"Đầu vào là file dữ liệu môn học phân bổ theo ngành và theo khóa.{tpl_link('Mon_NganhHoc_Template.xlsx')}"
}

col_info, col_action = st.columns([1.1, 1], gap="medium")

with col_info:
    st.markdown(f"""
        <div class="instruction-card">
            <strong>📌 YÊU CẦU DỮ LIỆU ĐẦU VÀO:</strong><br><br>
            {instructions.get(choice, "Vui lòng upload các file Excel theo đúng định dạng.")}
        </div>
    """, unsafe_allow_html=True)
    
    keywords_str = ""
    has_header = True
    msv_file = None
    
    if choice == "Điền KHLM (Updated) (*)":
        keywords_str = st.text_input("Thiết lập điều kiện (Từ khóa địa phương):", value="DNP, ĐN ( học tại HCM)")
    elif choice == "Kiểm tra điều kiện đăng ký môn học (Cả lớp)":
        has_header = st.checkbox("Tùy chọn: File chứa dòng tiêu đề (Header row)", value=True)
    elif choice in ["Gom điểm UNI", "Gom điểm"]:
        msv_file = st.file_uploader("📥 Tải lên Danh sách Mã Cần Tra (.xlsx)", type=['xlsx', 'xls', 'xlsb', 'xlsm'])

    status_container = st.empty() 

with col_action:
    uploader_label = "Kéo thả các file Dữ liệu Nguồn vào đây" if choice in ["Gom điểm UNI", "Gom điểm"] else "Kéo thả các file Excel vào đây"
    uploaded_files = st.file_uploader(uploader_label, accept_multiple_files=True, type=['xlsx', 'xlsb', 'xls', 'xlsm'])
    
    if st.button("🚀 XỬ LÝ DỮ LIỆU"):
        if not uploaded_files:
            st.error("⚠️ Vui lòng tải dữ liệu nguồn lên trước!")
        elif choice in ["Gom điểm UNI", "Gom điểm"] and not msv_file:
            st.error("⚠️ Bạn cần tải lên File Danh sách Mã Sinh Viên trước khi quét vét cạn.")
        else:
            temp_dir = tempfile.mkdtemp()
            try:
                if choice in ["Gom điểm UNI", "Gom điểm"]:
                    msv_path = os.path.join(temp_dir, "MSV_List.xlsx")
                    with open(msv_path, "wb") as f:
                        f.write(msv_file.getbuffer())
                    
                    data_dir = os.path.join(temp_dir, "data_source")
                    os.makedirs(data_dir, exist_ok=True)
                    for uf in uploaded_files:
                        with open(os.path.join(data_dir, uf.name), "wb") as f:
                            f.write(uf.getbuffer())
                else:
                    for uploaded_file in uploaded_files:
                        with open(os.path.join(temp_dir, uploaded_file.name), "wb") as f:
                            f.write(uploaded_file.getbuffer())
                
                with st.spinner('⚙️ Hệ thống đang xử lý, vui lòng giữ trang...'):
                    result_file = None
                    if choice == "Gộp File Nguồn": result_file = process_files_logic(temp_dir)
                    elif choice == "Tra cứu điểm sinh viên (Cả lớp)": result_file = scrape_ehou_logic(temp_dir, status_container)
                    elif choice == "Kiểm tra điều kiện đăng ký môn học (Cả lớp)": result_file = check_dk_dangky_logic(temp_dir, has_header)
                    elif choice == "Kiểm tra tiên quyết": result_file = check_tien_quyet_logic(temp_dir)
                    elif choice == "Gom điểm UNI": result_file = gom_diem_uni_logic(msv_path, data_dir)
                    elif choice == "Gom điểm": result_file = gom_diem_logic(msv_path, data_dir)
                    elif choice == "Điền KHLM (Updated) (*)": result_file = fill_khlm_logic(temp_dir, keywords_str)
                    elif choice == "Kiểm tra KHLM": result_file = check_khlm_logic(temp_dir)
                    elif choice == "Lọc kết quả học tập của sinh viên": result_file = compare_data_logic(temp_dir)
                    elif choice == "Lọc sinh viên học lại & học cải thiện": result_file = filter_sv_logic(temp_dir)
                    elif choice == "Xuất Mã lớp theo GK300": result_file = extract_class_names_logic(temp_dir)
                    elif choice == "Xuất KHHT theo GK300 (1*)": result_file = export_khht_logic(temp_dir)
                    elif choice == "Xuất DSSV theo GK300 (2*)": result_file = export_dssv_logic(temp_dir)
                    elif choice == "Xuất KHHTCT theo GK300 (3*)": result_file = export_khhtct_logic(temp_dir)
                    elif choice == "Thống kê số lượng theo lớp/nhóm lớp & môn (*)": result_file = fill_sllm_logic(temp_dir)
                    elif choice == "Xuất môn theo ngành học (*)": result_file = extract_courses_logic(temp_dir)

                    if result_file and os.path.exists(result_file):
                        status_container.empty()
                        with open(result_file, "rb") as f:
                            file_data = f.read()
                        
                        file_ext = os.path.splitext(result_file)[1].lower()
                        mime_type = "application/zip" if file_ext == '.zip' else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        
                        st.download_button(
                            label="⬇️ TẢI FILE KẾT QUẢ VỀ MÁY",
                            data=file_data,
                            file_name=os.path.basename(result_file),
                            mime=mime_type
                        )
                    else:
                        st.error("❌ Xử lý thất bại. Kiểm tra cấu trúc file.")
                        
            except Exception as e:
                status_container.empty()
                st.error(f"❌ Lỗi hệ thống: `{str(e)}`")
            finally:
                shutil.rmtree(temp_dir)
